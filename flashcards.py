from openpyxl import load_workbook
import pandas as pd
from colorama import Fore, Style
import random
import os
import tempfile
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

def select(sheets):
    os.system('cls')
    while True:
        n = input(f"How many words do you want to practice? (x = all)> ").strip()
        if n == "q":
            exit()
        if n:
           break   
    os.system('cls')
    for sheet,i in zip(sheets,range(len(sheets))):
        print(f"[{i+1}] - {sheet}")
    while True:
        opt = input(f"What do you want to practice? > ").strip()
        if opt == "q":
            exit()
        if opt and n == "x":
            return n,int(opt)
        if opt:
            return int(n),int(opt)

def write(filename, what, sheet, val):
    wb = load_workbook(filename)
    ws = wb[sheet]
    table = ws._tables[sheet]
    rng = table.ref
    sc,ec = rng.split(":")
    sr = ws[sc].row
    er = ws[ec].row
    headers = {cell.value: cell.column for cell in ws[sr]}
    wc = headers["Znam"]
    for row in ws.iter_rows(min_row=sr+ 1, max_row=er, values_only=False):
        for cell in row:
            if cell.value and what in str(cell.value).lower():
                ws.cell(row=cell.row, column=wc, value=val)
    wb.save(filename)
    wb.close()

def play(obj, filename, sheet):
    os.system('cls')
    unused = obj
    n = len(obj)
    used = []
    to_learn = []
    while unused:
        entry = random.choice(unused)
        used.append(entry)
        inpt = input(f"\n{Fore.GREEN}Riječ {len(used)}/{n}: {Fore.BLUE}{entry[-1]} {Style.RESET_ALL}> ")
        if inpt == "q":
            break    
        strng = " - ".join(map(str, entry[1:-1])) 
        print(f"{Fore.YELLOW}{strng}{Style.RESET_ALL}")
        inpt = input(f"Did you know? > ")
        if inpt == "x":
            to_learn.append(entry)
            if sheet:
                write(filename, entry[-1], sheet, "x")
        else:
            if sheet:
                write(filename, entry[-1], sheet, "da")
        unused.remove(entry)
        if not unused:
            pdfmetrics.registerFont(TTFont('Arial', 'Arial.ttf'))
            score = n-len(to_learn)
            print(f"\n{Fore.YELLOW}Score: {score}/{n} ({(score/n)*100})%{Style.RESET_ALL}")
            if score != n:
                with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp:
                    c = canvas.Canvas(tmp.name, pagesize=letter)
                    c.setFont("Arial", 12)
                    y = 750
                    line = 0
                    for i,item in zip(range(len(to_learn)),to_learn):
                        c.drawString(100,y,f"{i}. " + str(item))
                        y -= 20
                        line += 1
                        if line >= 35:
                            c.showPage()
                            c.setFont("Arial", 12)
                            y = 750
                            line = 0
                    c.save()
                os.startfile(tmp.name)
            input(f"\nPress enter to leave...")
            break

data = []
filename = r"C:\Users\User\OneDrive\Belgeler\deutsch.xlsx"
file = pd.ExcelFile(filename)
sheets = file.sheet_names
condition = ""
o = input(f"Do you want hard difficulty? > ")
if o == "y":
    condition = "da"
print(f"Fetching data...")
for sheet in sheets:
    df = pd.read_excel(filename, sheet_name=sheet, header=1)
    words = []
    for index,row in df.iterrows():
        temp = []
        for column_name, value in row.items():
            if "Unnamed" not in column_name:
                if value == "sl" or value == condition:
                    break
                temp.append(str(value).strip())
        if temp:
            words.append(temp)
    data.append(words)

while True:
    nwords,opt = select(sheets)
    if nwords == "x":
        play(data[opt-1],filename,sheets[opt-1])
    else:
        play(data[opt-1][:nwords],filename,sheets[opt-1])